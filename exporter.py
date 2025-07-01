# exporter.py-Excel 样式与导出

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment


def export_to_excel(df, output_file="英语教授课程安排表.xlsx"):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='课程安排')
        worksheet = writer.sheets['课程安排']

        time_col_index = df.columns.get_loc("时间") + 1
        for row in worksheet.iter_rows(min_col=time_col_index, max_col=time_col_index):
            for cell in row:
                cell.number_format = '@'

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value).encode('utf-8'))
                        max_length = max(max_length, length)
                except:
                    pass
            adjusted_width = (max_length * 1.2) / 1.1
            worksheet.column_dimensions[column].width = min(max(adjusted_width, 15), 50)

        header_fill = PatternFill(start_color='4F94CD', end_color='4F94CD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment

        date_col_index = df.columns.get_loc("日期") + 1
        from config import WEEKDAY_COLOR_MAP
        for row in worksheet.iter_rows(min_row=2):
            cell = row[date_col_index - 1]
            weekday = cell.value
            color_hex = WEEKDAY_COLOR_MAP.get(weekday)
            if color_hex:
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                cell.fill = fill

    print(f"✅ 课程安排表已成功导出到：{output_file}")